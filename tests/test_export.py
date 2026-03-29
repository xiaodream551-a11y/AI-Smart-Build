# -*- coding: utf-8 -*-

import json

from openpyxl import load_workbook

from engine.export import export_model_data, export_to_excel, export_to_json
from tools.offline_runtime import (
    FakeBuiltInCategory,
    FakeCurve,
    FakeCurveLocation,
    FakeDocument,
    FakeElement,
    FakeElementType,
    FakeParameter,
    FakePointLocation,
    FakeStorageType,
    FakeXYZ,
    make_story_levels,
)
from pyrevit import DB


def test_export_model_data_collects_columns_beams_and_slabs():
    levels = make_story_levels(2)
    column_type = FakeElementType(
        9001,
        lookup_params={
            "b": FakeParameter(500.0 / 304.8, storage_type=FakeStorageType.Double),
            "h": FakeParameter(500.0 / 304.8, storage_type=FakeStorageType.Double),
        },
    )
    beam_type = FakeElementType(9002, name="300x600")

    column = FakeElement(
        101,
        FakeBuiltInCategory.OST_StructuralColumns,
        level_id=1,
        params={
            DB.BuiltInParameter.FAMILY_TOP_LEVEL_PARAM: FakeParameter(levels[1].Id, storage_type=FakeStorageType.ElementId),
        },
        location=FakePointLocation(FakeXYZ(6000.0 / 304.8, 0.0, 0.0)),
        symbol=column_type,
        name="柱",
    )
    beam = FakeElement(
        102,
        FakeBuiltInCategory.OST_StructuralFraming,
        level_id=2,
        location=FakeCurveLocation(FakeCurve(
            FakeXYZ(0.0, 0.0, levels[1].Elevation),
            FakeXYZ(6000.0 / 304.8, 0.0, levels[1].Elevation),
        )),
        symbol=beam_type,
        name="梁",
    )
    slab = FakeElement(
        103,
        FakeBuiltInCategory.OST_Floors,
        level_id=2,
        params={
            DB.BuiltInParameter.HOST_AREA_COMPUTED: FakeParameter(
                36.0 / (0.3048 ** 2),
                storage_type=FakeStorageType.Double,
            ),
        },
        name="板",
    )

    doc = FakeDocument(levels=levels, elements=[column, beam, slab], element_types=[column_type, beam_type])

    data = export_model_data(doc)

    assert [item["name"] for item in data["levels"]] == ["±0.000", "F1", "屋面"]
    assert data["columns"][0]["x_mm"] == 6000.0
    assert data["columns"][0]["base_level"] == "±0.000"
    assert data["columns"][0]["top_level"] == "F1"
    assert data["columns"][0]["section"] == "500x500"
    assert data["beams"][0]["end_x_mm"] == 6000.0
    assert data["beams"][0]["level"] == "F1"
    assert data["beams"][0]["section"] == "300x600"
    assert data["slabs"][0]["area_sqm"] == 36.0
    assert data["summary"] == {"columns": 1, "beams": 1, "slabs": 1}


def test_export_to_json_writes_file(tmp_path):
    filepath = tmp_path / "model.json"
    data = {
        "columns": [{"id": 1}],
        "beams": [],
        "slabs": [],
        "summary": {"columns": 1, "beams": 0, "slabs": 0},
        "levels": [],
    }

    export_to_json(data, str(filepath))

    loaded = json.loads(filepath.read_text(encoding="utf-8"))
    assert loaded["columns"][0]["id"] == 1
    assert loaded["summary"]["columns"] == 1


def test_export_to_excel_writes_expected_sheets(tmp_path):
    filepath = tmp_path / "model.xlsx"
    data = {
        "columns": [
            {"id": 1, "x_mm": 0, "y_mm": 0, "base_level": "±0.000", "top_level": "F1", "section": "500x500"},
        ],
        "beams": [
            {"id": 2, "start_x_mm": 0, "start_y_mm": 0, "end_x_mm": 6000, "end_y_mm": 0, "level": "F1", "section": "300x600"},
        ],
        "slabs": [
            {"id": 3, "level": "F1", "area_sqm": 36.0},
        ],
        "summary": {"columns": 1, "beams": 1, "slabs": 1},
        "levels": [],
    }

    export_to_excel(data, str(filepath))

    workbook = load_workbook(str(filepath))
    assert workbook.sheetnames == ["柱", "梁", "板", "汇总"]
    assert workbook["柱"]["A2"].value == 1
    assert workbook["柱"]["F2"].value == "500x500"
    assert workbook["梁"]["G2"].value == "300x600"
    assert workbook["板"]["C2"].value == 36.0
    assert workbook["汇总"]["B2"].value == 1
