# -*- coding: utf-8 -*-
"""生成 Excel 构件导入模板"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADERS = [u"类型", u"X(mm)", u"Y(mm)", u"楼层", u"截面"]

SAMPLE_ROWS = [
    [u"柱", 0, 0, 1, u"500x500"],
    [u"柱", 6000, 0, 1, u"500x500"],
    [u"柱", 12000, 0, 1, u"500x500"],
    [u"梁", u"0,0", u"6000,0", 2, u"300x600"],
    [u"梁", u"6000,0", u"12000,0", 2, u"300x600"],
    [u"梁", u"12000,0", u"18000,0", 2, u"300x600"],
]


def _display_width(value):
    text = u"" if value is None else u"{}".format(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


def _set_column_widths(worksheet):
    for column_cells in worksheet.columns:
        max_width = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            max_width = max(max_width, _display_width(cell.value))

        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = max_width + 2


def build_template(output_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = u"构件清单"

    header_font = Font(bold=True)
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9D9D9"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    worksheet.append(HEADERS)
    for row in SAMPLE_ROWS:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    _set_column_widths(worksheet)

    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    workbook.save(output_path)


def main():
    project_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(
        project_dir,
        u"AI智建.extension",
        u"templates",
        u"构件导入模板.xlsx"
    )
    build_template(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
