# -*- coding: utf-8 -*-
"""从 Excel 表格批量导入构件"""

__doc__ = "选择 Excel 构件清单，批量创建柱、梁"
__title__ = "Excel\n导入"
__author__ = "AI智建"

import openpyxl

from pyrevit import revit, DB, forms, script

from engine.column import create_column
from engine.beam import create_beam

try:
    string_types = (basestring,)
    text_type = unicode
except NameError:
    string_types = (str,)
    text_type = str


REQUIRED_HEADERS = {
    "type": u"类型",
    "x": u"X(mm)",
    "y": u"Y(mm)",
    "floor": u"楼层",
    "section": u"截面",
}


def _to_text(value):
    if value is None:
        return u""
    if isinstance(value, string_types):
        return value.strip()
    return text_type(value).strip()


def _normalize_header(value):
    return _to_text(value).replace(" ", "").replace(u"\u3000", "").lower()


def _normalize_type(value):
    return _to_text(value).replace(" ", "").lower()


def _normalize_number_text(value):
    return _to_text(value).replace(" ", "")


def _parse_float(value, field_name):
    text = _normalize_number_text(value)
    if not text:
        raise ValueError(u"{}为空".format(field_name))
    if "," in text or u"\uff0c" in text:
        raise ValueError(u"{}应为单个数值".format(field_name))
    return float(text)


def _parse_floor(value):
    floor = int(float(_normalize_number_text(value)))
    if floor <= 0:
        raise ValueError(u"楼层必须大于 0")
    return floor


def _parse_point(value, field_name):
    text = _to_text(value)
    if not text:
        raise ValueError(u"{}为空".format(field_name))

    normalized = text.replace(u"\uff0c", ",").replace(";", ",").replace(" ", "")
    parts = [part for part in normalized.split(",") if part]
    if len(parts) != 2:
        raise ValueError(u"{}格式错误，应为 x,y".format(field_name))

    return float(parts[0]), float(parts[1])


def _get_cell(row, index):
    if index >= len(row):
        return None
    return row[index]


def _collect_levels(doc):
    levels = list(DB.FilteredElementCollector(doc).OfClass(DB.Level))
    levels.sort(key=lambda level: level.Elevation)
    return levels


def _parse_headers(ws):
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError(u"Excel 缺少表头")

    header_map = {}
    for index, value in enumerate(header_row):
        normalized = _normalize_header(value)
        if normalized:
            header_map[normalized] = index

    missing = []
    result = {}
    for key, name in REQUIRED_HEADERS.items():
        header_index = header_map.get(_normalize_header(name))
        if header_index is None:
            missing.append(name)
        else:
            result[key] = header_index

    if missing:
        raise ValueError(u"Excel 缺少列：{}".format(u"，".join(missing)))

    return result


def _parse_row(row_index, row, header_map, levels):
    if not row:
        raise ValueError(u"空行")

    if all(_to_text(value) == u"" for value in row):
        raise ValueError(u"空行")

    type_value = _normalize_type(_get_cell(row, header_map["type"]))
    floor = _parse_floor(_get_cell(row, header_map["floor"]))
    section = _to_text(_get_cell(row, header_map["section"]))

    if not section:
        raise ValueError(u"截面为空")

    if type_value in (u"柱", "column", "columns"):
        if floor >= len(levels):
            raise ValueError(u"楼层 {} 超出当前标高范围".format(floor))

        return {
            "row": row_index,
            "kind": "column",
            "x": _parse_float(_get_cell(row, header_map["x"]), u"X(mm)"),
            "y": _parse_float(_get_cell(row, header_map["y"]), u"Y(mm)"),
            "base_level": levels[floor - 1],
            "top_level": levels[floor],
            "section": section,
        }

    if type_value in (u"梁", "beam", "beams"):
        if floor >= len(levels):
            raise ValueError(u"楼层 {} 超出当前标高范围".format(floor))

        start_x, start_y = _parse_point(_get_cell(row, header_map["x"]), u"X(mm)")
        end_x, end_y = _parse_point(_get_cell(row, header_map["y"]), u"Y(mm)")
        return {
            "row": row_index,
            "kind": "beam",
            "start_x": start_x,
            "start_y": start_y,
            "end_x": end_x,
            "end_y": end_y,
            "level": levels[floor],
            "section": section,
        }

    raise ValueError(u"不支持的类型: {}".format(type_value or u"<空>"))


def _log_skip(output, logger, row_index, reason, skipped_logs):
    message = u"第 {} 行已跳过：{}".format(row_index, reason)
    skipped_logs.append(message)
    logger.warning(message)


def main():
    doc = revit.doc
    output = script.get_output()
    logger = script.get_logger()

    file_path = forms.pick_file(file_ext='xlsx')
    if not file_path:
        script.exit()

    levels = _collect_levels(doc)
    if len(levels) < 2:
        forms.alert(u"当前模型标高不足，至少需要 2 个标高。", title=u"AI 智建")
        script.exit()

    output.print_md(u"## AI 智建 — Excel 导入")
    output.print_md(u"- 文件：`{}`".format(file_path))

    workbook = None
    operations = []
    skipped_logs = []
    skipped_count = 0

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        header_map = _parse_headers(worksheet)

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):
            try:
                operations.append(_parse_row(row_index, row, header_map, levels))
            except Exception as err:
                skipped_count += 1
                _log_skip(output, logger, row_index, _to_text(err), skipped_logs)
    except Exception as err:
        forms.alert(u"Excel 读取失败：{}".format(_to_text(err)), title=u"AI 智建")
        script.exit()
    finally:
        if workbook:
            workbook.close()

    success_count = 0

    with revit.Transaction(u"AI智建：Excel导入"):
        for item in operations:
            try:
                if item["kind"] == "column":
                    create_column(
                        doc,
                        item["x"], item["y"],
                        item["base_level"], item["top_level"],
                        item["section"]
                    )
                elif item["kind"] == "beam":
                    create_beam(
                        doc,
                        item["start_x"], item["start_y"],
                        item["end_x"], item["end_y"],
                        item["level"], item["section"]
                    )
                else:
                    raise ValueError(u"不支持的构件类型: {}".format(item["kind"]))

                success_count += 1
            except Exception as err:
                skipped_count += 1
                _log_skip(output, logger, item["row"], _to_text(err), skipped_logs)

    if skipped_logs:
        output.print_md(u"### 跳过记录")
        for message in skipped_logs:
            output.print_md(u"- {}".format(message))

    forms.alert(
        u"成功导入 {} 个构件，跳过 {} 行".format(success_count, skipped_count),
        title=u"AI 智建"
    )


if __name__ == "__main__":
    main()
