# -*- coding: utf-8 -*-
"""Model data export."""

import io
import json
import os

from pyrevit import DB

from engine.element_utils import (
    get_element_area_sqm,
    get_element_level_id,
    get_element_section_text,
    get_level_id_from_parameter,
    is_valid_element_id,
    resolve_level_name,
)
from utils import feet_to_mm, get_sorted_levels, _get_name


_CATEGORY_TO_NAME = {
    DB.BuiltInCategory.OST_StructuralColumns: "columns",
    DB.BuiltInCategory.OST_StructuralFraming: "beams",
    DB.BuiltInCategory.OST_Floors: "slabs",
}


def export_model_data(doc):
    levels = get_sorted_levels(doc)
    data = {
        "levels": [
            {
                "name": _get_name(level),
                "elevation_mm": _round_mm(feet_to_mm(level.Elevation)),
            }
            for level in levels
        ],
        "columns": [],
        "beams": [],
        "slabs": [],
        "summary": {
            "columns": 0,
            "beams": 0,
            "slabs": 0,
        },
    }

    collectors = (
        (DB.BuiltInCategory.OST_StructuralColumns, _collect_column_data),
        (DB.BuiltInCategory.OST_StructuralFraming, _collect_beam_data),
        (DB.BuiltInCategory.OST_Floors, _collect_slab_data),
    )
    for category, builder in collectors:
        items = []
        for element in DB.FilteredElementCollector(doc).OfCategory(category).WhereElementIsNotElementType():
            item = builder(doc, element)
            if item:
                items.append(item)
        key = _CATEGORY_TO_NAME[category]
        data[key] = items
        data["summary"][key] = len(items)

    return data


def export_to_json(data, filepath):
    _ensure_parent_dir(filepath)
    with io.open(filepath, "w", encoding="utf-8") as output_file:
        json.dump(data, output_file, ensure_ascii=False, indent=2)


def _export_to_csv(data, filepath):
    """Fallback: export as CSV when openpyxl fails in IronPython."""
    import csv
    csv_path = filepath.replace(".xlsx", ".csv")
    with io.open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([u"类型", u"ID", u"X(mm)", u"Y(mm)", u"底标高/起点X",
                         u"顶标高/起点Y", u"终点X", u"终点Y", u"标高", u"截面", u"面积(m²)"])
        for item in data.get("columns", []):
            writer.writerow([u"柱", item["id"], item["x_mm"], item["y_mm"],
                             item["base_level"], item["top_level"], "", "", "", item["section"], ""])
        for item in data.get("beams", []):
            writer.writerow([u"梁", item["id"], "", "", item["start_x_mm"], item["start_y_mm"],
                             item["end_x_mm"], item["end_y_mm"], item["level"], item["section"], ""])
        for item in data.get("slabs", []):
            writer.writerow([u"板", item["id"], "", "", "", "", "", "", item["level"], "", item["area_sqm"]])
    return csv_path


def export_to_excel(data, filepath):
    try:
        from openpyxl import Workbook

        _ensure_parent_dir(filepath)

        workbook = Workbook(write_only=True)

        def _write_sheet(wb, title, headers, rows):
            ws = wb.create_sheet(title=title)
            ws.append(headers)
            for row in rows:
                ws.append(row)

        _write_sheet(workbook, u"柱", [u"ID", u"X(mm)", u"Y(mm)", u"底标高", u"顶标高", u"截面"], [
            [item["id"], item["x_mm"], item["y_mm"],
             item["base_level"], item["top_level"], item["section"]]
            for item in data.get("columns", [])
        ])
        _write_sheet(workbook, u"梁", [u"ID", u"起点X", u"起点Y", u"终点X", u"终点Y", u"标高", u"截面"], [
            [item["id"], item["start_x_mm"], item["start_y_mm"],
             item["end_x_mm"], item["end_y_mm"], item["level"], item["section"]]
            for item in data.get("beams", [])
        ])
        _write_sheet(workbook, u"板", [u"ID", u"标高", u"面积(m²)"], [
            [item["id"], item["level"], item["area_sqm"]]
            for item in data.get("slabs", [])
        ])
        _write_sheet(workbook, u"汇总", [u"项目", u"数量"], [
            [u"柱", data.get("summary", {}).get("columns", 0)],
            [u"梁", data.get("summary", {}).get("beams", 0)],
            [u"板", data.get("summary", {}).get("slabs", 0)],
        ])

        workbook.save(filepath)
    except Exception:
        csv_path = _export_to_csv(data, filepath)
        raise Exception(u"Excel 不可用，已导出 CSV: " + csv_path)


def _append_sheet(workbook, title, headers, rows):
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    return worksheet


def _collect_column_data(doc, element):
    point = getattr(getattr(element, "Location", None), "Point", None)
    if point is None:
        return None

    base_level = resolve_level_name(doc, get_element_level_id(DB, element))
    top_level = resolve_level_name(
        doc,
        get_level_id_from_parameter(DB, element, DB.BuiltInParameter.FAMILY_TOP_LEVEL_PARAM)
    )
    return {
        "id": element.Id.IntegerValue,
        "x_mm": _round_mm(feet_to_mm(point.X)),
        "y_mm": _round_mm(feet_to_mm(point.Y)),
        "base_level": base_level,
        "top_level": top_level,
        "section": get_element_section_text(DB, element),
    }


def _collect_beam_data(doc, element):
    curve = getattr(getattr(element, "Location", None), "Curve", None)
    if curve is None:
        return None

    start = curve.GetEndPoint(0)
    end = curve.GetEndPoint(1)
    return {
        "id": element.Id.IntegerValue,
        "start_x_mm": _round_mm(feet_to_mm(start.X)),
        "start_y_mm": _round_mm(feet_to_mm(start.Y)),
        "end_x_mm": _round_mm(feet_to_mm(end.X)),
        "end_y_mm": _round_mm(feet_to_mm(end.Y)),
        "level": resolve_level_name(doc, get_element_level_id(DB, element)),
        "section": get_element_section_text(DB, element),
    }


def _collect_slab_data(doc, element):
    area_sqm = round(get_element_area_sqm(DB, element), 3)
    return {
        "id": element.Id.IntegerValue,
        "level": resolve_level_name(doc, get_element_level_id(DB, element)),
        "area_sqm": area_sqm,
    }


def _round_mm(value):
    return round(float(value), 3)


def _ensure_parent_dir(filepath):
    parent_dir = os.path.dirname(filepath)
    if parent_dir and not os.path.exists(parent_dir):
        os.makedirs(parent_dir)
